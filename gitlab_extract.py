"""
SCRIPT 2 of 3 — GitLab extraction into the shared master Excel file.

Reads harness_gitlab_master.xlsx (as populated by Script 1). For every row:

  1. Resolve the GitLab project ("Corresponding Gitlab component"):
       - If that column is ALREADY filled in — by hand, or by a previous
         run of this script — it's trusted as-is and used directly, no
         searching. This is what makes re-runs fast and stable: fix an
         ambiguous match once, and every run after that just uses it.
       - Otherwise, searches GitLab using a few candidate terms (the
         Harness service name, and the last path segment of Artifact Path,
         which often tracks the real repo name more closely than the
         Harness display name does). Only auto-fills the match when it's
         unambiguous. Anything ambiguous or not found is left BLANK, with
         the candidates it considered noted in Comments — never guessed.

  2. Find the yml file holding the target variables. Tries GitLab's code
     search first (searches file CONTENTS for the variable names, so it
     doesn't need to know the folder structure up front — useful given
     how inconsistent that structure is across repos), then falls back to
     a short list of common file paths if search comes back empty.

  3. Fill in New_Branch (GitLab's actual default branch, or the existing
     Harness branch if that can't be determined — and skipped entirely
     for services whose manifest isn't Git-sourced at all), New_ssc_appname
     and New_ssc_appversion per VARIABLE_MAPPING below, plus which file
     they came from.

Run Script 1 first. Script 3 (Harness update) reads this same file next.

pip install requests pyyaml openpyxl --break-system-packages
"""

import os
import sys
import time
import base64
import logging
from datetime import datetime

import yaml
import requests
from openpyxl import load_workbook

from harness_config import load_config, require_gitlab

# ===========================================================================
# CONFIG
# ===========================================================================

_config = load_config()
_gitlab = require_gitlab(_config)

GITLAB_BASE_URL = _gitlab["base_url"].rstrip("/")
GITLAB_PRIVATE_TOKEN = _gitlab["private_token"]
GITLAB_BRANCH = _gitlab.get("branch") or "main"
GITLAB_GROUP_PATH = _gitlab.get("group_path")  # optional — narrows search scope, e.g. "fiserv/backend-services"

MASTER_EXCEL_FILE = "harness_gitlab_master.xlsx"
LOG_FILE = f"script2_gitlab_extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

# Harness variable name -> GitLab yml key name (confirmed from the client's note)
VARIABLE_MAPPING = {
    "ssc_appname": "FORTIFY_APPLICATION_VERSION",
    "ssc_appversion": "FORTIFY_APPLICATION_NAME",
}
NEW_COLUMN_FOR_HARNESS_VAR = {
    "ssc_appname": "New_ssc_appname",
    "ssc_appversion": "New_ssc_appversion",
}

# Tried in order only if GitLab's code search comes back empty
YML_FILE_CANDIDATES = [
    "values.yaml", "values.yml",
    "config/values.yaml", "helm/values.yaml",
]

SLEEP_BETWEEN_ROWS = 0.3
MAX_RETRIES = 3
RETRY_BACKOFF_SECONDS = 2

# ===========================================================================
# Shared column schema — MUST exactly match Script 1's ALL_COLUMNS.
# ===========================================================================

ALL_COLUMNS = [
    "Service Identifier", "Service Name", "Service Url",
    "Corresponding Gitlab component", "GitLab File Path",
    "Existing_Branch", "New_Branch",
    "Existing_ssc_appname", "New_ssc_appname",
    "Existing_ssc_appversion", "New_ssc_appversion",
    "Artifact Path",
    "Approved for Update (Y/N)", "Variables Updated",
    "Comments", "Last Run Timestamp",
]

OWNED_COLUMNS = [
    "Corresponding Gitlab component", "GitLab File Path",
    "New_Branch", "New_ssc_appname", "New_ssc_appversion",
]

# ===========================================================================
# Logging
# ===========================================================================

logger = logging.getLogger("script2")
logger.setLevel(logging.DEBUG)
fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
logger.addHandler(fh)
logger.addHandler(ch)

# ===========================================================================
# GitLab API helpers
# ===========================================================================

GITLAB_HEADERS = {"PRIVATE-TOKEN": GITLAB_PRIVATE_TOKEN}


def request_with_retry(method: str, url: str, **kwargs) -> requests.Response:
    last_exc = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.request(method, url, timeout=30, **kwargs)
            if resp.status_code >= 500:
                raise requests.HTTPError(f"{resp.status_code} server error")
            return resp
        except requests.RequestException as exc:
            last_exc = exc
            logger.warning("Request failed (attempt %s/%s) %s %s: %s",
                            attempt, MAX_RETRIES, method, url, exc)
            time.sleep(RETRY_BACKOFF_SECONDS * attempt)
    raise last_exc


def search_gitlab_projects(term: str) -> list[dict]:
    params = {"search": term}
    if GITLAB_GROUP_PATH:
        # GitLab's group-projects endpoint accepts either a numeric group ID
        # or a URL-encoded full path (e.g. "fiserv/backend-services") in
        # this slot — quote() turns the "/" into %2F, which is what makes
        # the path form work here.
        url = f"{GITLAB_BASE_URL}/api/v4/groups/{requests.utils.quote(GITLAB_GROUP_PATH, safe='')}/projects"
    else:
        url = f"{GITLAB_BASE_URL}/api/v4/projects"
    resp = request_with_retry("GET", url, headers=GITLAB_HEADERS, params=params)
    resp.raise_for_status()
    return resp.json()


def get_project_by_path(path: str) -> dict | None:
    url = f"{GITLAB_BASE_URL}/api/v4/projects/{requests.utils.quote(path, safe='')}"
    resp = request_with_retry("GET", url, headers=GITLAB_HEADERS)
    if resp.status_code == 200:
        return resp.json()
    return None


def resolve_gitlab_project(existing_component: str, service_name: str, artifact_path: str):
    """
    Returns (project_dict_or_None, note). note is "" on a clean resolution;
    otherwise it explains what happened, for the Comments column.
    """
    if existing_component:
        project = get_project_by_path(existing_component)
        if project:
            return project, ""
        return None, (f"'{existing_component}' (already set) was not found on "
                       f"GitLab — left as-is, please check it")

    terms, seen = [], set()
    for term in [service_name, (artifact_path or "").rstrip("/").split("/")[-1]]:
        if term and term not in seen:
            seen.add(term)
            terms.append(term)

    pool = {}
    for term in terms:
        try:
            for proj in search_gitlab_projects(term):
                pool[proj["id"]] = proj
        except requests.HTTPError as exc:
            logger.warning("GitLab search failed for '%s': %s", term, exc)

    if not pool:
        return None, f"No GitLab project found for candidates: {terms}"

    lower_terms = [t.lower() for t in terms]
    exact_matches = [p for p in pool.values() if p.get("name", "").lower() in lower_terms]
    if len(exact_matches) == 1:
        return exact_matches[0], ""
    if len(pool) == 1:
        return next(iter(pool.values())), ""

    labels = [p.get("path_with_namespace", p.get("name")) for p in pool.values()]
    return None, (f"Ambiguous match for candidates {terms} -> found {labels}; "
                  f"set 'Corresponding Gitlab component' manually to resolve")


def code_search_yml_file(project_id: int, search_term: str) -> str | None:
    url = f"{GITLAB_BASE_URL}/api/v4/projects/{project_id}/search"
    params = {"scope": "blobs", "search": search_term, "ref": GITLAB_BRANCH}
    resp = request_with_retry("GET", url, headers=GITLAB_HEADERS, params=params)
    if resp.status_code != 200:
        return None
    results = resp.json()
    if not results:
        return None
    yml_hits = [r for r in results if r.get("path", "").lower().endswith((".yml", ".yaml"))]
    chosen = yml_hits[0] if yml_hits else results[0]
    return chosen.get("path")


def read_gitlab_file(project_id: int, file_path: str) -> dict | None:
    encoded_path = requests.utils.quote(file_path, safe="")
    url = f"{GITLAB_BASE_URL}/api/v4/projects/{project_id}/repository/files/{encoded_path}"
    resp = request_with_retry("GET", url, headers=GITLAB_HEADERS, params={"ref": GITLAB_BRANCH})
    if resp.status_code != 200:
        return None
    content_b64 = resp.json()["content"]
    content = base64.b64decode(content_b64).decode("utf-8", errors="replace")
    try:
        return yaml.safe_load(content) or {}
    except yaml.YAMLError as exc:
        logger.warning("YAML parse error in %s: %s", file_path, exc)
        return None


def find_and_read_yaml_file(project_id: int, target_keys: list[str]) -> tuple[str, dict] | tuple[None, None]:
    for key in target_keys:
        path = code_search_yml_file(project_id, key)
        if path:
            parsed = read_gitlab_file(project_id, path)
            if parsed:
                return path, parsed
    for path in YML_FILE_CANDIDATES:
        parsed = read_gitlab_file(project_id, path)
        if parsed:
            return path, parsed
    return None, None


def extract_vars_from_yaml(parsed_yaml: dict, var_names: list[str]) -> dict:
    """Recursive lookup — checked at every level regardless of nesting depth."""
    found = {}

    def _search(node):
        if isinstance(node, dict):
            for k, v in node.items():
                if k in var_names and k not in found:
                    found[k] = v
                else:
                    _search(v)
        elif isinstance(node, list):
            for item in node:
                _search(item)

    _search(parsed_yaml)
    return found


# ===========================================================================
# Master Excel file — generic read/write helpers
# ===========================================================================


def get_cell(ws, row_idx: int, column_name: str):
    col_idx = ALL_COLUMNS.index(column_name) + 1
    return ws.cell(row=row_idx, column=col_idx).value


def set_owned_cell(ws, row_idx: int, column_name: str, value):
    if column_name not in OWNED_COLUMNS:
        raise ValueError(f"Script 2 tried to write a column it doesn't own: {column_name}")
    col_idx = ALL_COLUMNS.index(column_name) + 1
    ws.cell(row=row_idx, column=col_idx, value=value)


def set_script_comment(ws, row_idx: int, script_tag: str, text: str):
    """Replaces this script's previous comment(s) on the row rather than
    stacking new ones on top; other scripts' comments are left untouched."""
    col_idx = ALL_COLUMNS.index("Comments") + 1
    cell = ws.cell(row=row_idx, column=col_idx)
    existing_lines = (cell.value or "").split("\n")
    other_lines = [line for line in existing_lines if line and not line.startswith(f"[{script_tag} ")]
    if text:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        other_lines.append(f"[{script_tag} {stamp}] {text}")
    cell.value = "\n".join(other_lines) if other_lines else None


def set_timestamp(ws, row_idx: int):
    col_idx = ALL_COLUMNS.index("Last Run Timestamp") + 1
    ws.cell(row=row_idx, column=col_idx, value=datetime.now().isoformat(timespec="seconds"))


# ===========================================================================
# Per-row processing
# ===========================================================================


def process_row(ws, row_idx: int) -> str:
    notes = []
    svc_name = get_cell(ws, row_idx, "Service Name")
    if not svc_name:
        set_script_comment(ws, row_idx, "Script2", "Skipped: no Service Name yet (run Script 1 first)")
        return "skipped"

    artifact_path = get_cell(ws, row_idx, "Artifact Path")
    existing_branch = get_cell(ws, row_idx, "Existing_Branch")
    existing_component = get_cell(ws, row_idx, "Corresponding Gitlab component")

    project, resolve_note = resolve_gitlab_project(existing_component, svc_name, artifact_path)
    if resolve_note:
        notes.append(resolve_note)

    if not project:
        set_script_comment(ws, row_idx, "Script2", "; ".join(notes))
        return "no_project"

    component_label = project.get("path_with_namespace", project.get("name"))
    if not existing_component:
        set_owned_cell(ws, row_idx, "Corresponding Gitlab component", component_label)

    if existing_branch == "Harness":
        notes.append("Manifest is Harness File Store (not Git) — branch check skipped")
    else:
        new_branch = project.get("default_branch") or existing_branch
        set_owned_cell(ws, row_idx, "New_Branch", new_branch)

    target_keys = list(VARIABLE_MAPPING.values())
    file_path, parsed_yaml = find_and_read_yaml_file(project["id"], target_keys)

    if not file_path:
        notes.append(f"No yml file found containing the target variables in {component_label}")
        set_script_comment(ws, row_idx, "Script2", "; ".join(notes))
        return "no_yaml"

    set_owned_cell(ws, row_idx, "GitLab File Path", file_path)
    gitlab_values = extract_vars_from_yaml(parsed_yaml, target_keys)

    for harness_var, gitlab_key in VARIABLE_MAPPING.items():
        if gitlab_key not in gitlab_values:
            notes.append(f"'{gitlab_key}' not found in {file_path}")
            continue
        set_owned_cell(ws, row_idx, NEW_COLUMN_FOR_HARNESS_VAR[harness_var], str(gitlab_values[gitlab_key]))

    set_script_comment(ws, row_idx, "Script2", "; ".join(notes))
    return "ok"


# ===========================================================================
# Main
# ===========================================================================


def main():
    logger.info("=== Script 2 (GitLab extract) started ===")

    if not os.path.exists(MASTER_EXCEL_FILE):
        sys.exit(f"{MASTER_EXCEL_FILE} not found — run Script 1 first.")

    wb = load_workbook(MASTER_EXCEL_FILE)
    ws = wb.active
    header = [c.value for c in ws[1]]
    if header != ALL_COLUMNS:
        sys.exit(
            f"{MASTER_EXCEL_FILE}'s header doesn't match the expected schema.\n"
            f"Expected: {ALL_COLUMNS}\nFound:    {header}"
        )

    total_rows = ws.max_row - 1
    counters = {}

    for row_idx in range(2, ws.max_row + 1):
        logger.info("[row %s/%s]", row_idx - 1, total_rows)
        try:
            result = process_row(ws, row_idx)
        except Exception as exc:  # noqa: BLE001 — keep going across all rows
            logger.error("Unhandled error on row %s: %s", row_idx, exc)
            set_script_comment(ws, row_idx, "Script2", f"ERROR: {exc}")
            result = "error"
        counters[result] = counters.get(result, 0) + 1
        set_timestamp(ws, row_idx)
        time.sleep(SLEEP_BETWEEN_ROWS)

    wb.save(MASTER_EXCEL_FILE)
    logger.info("=== Script 2 complete: %s ===", counters)
    print(f"\nDone. Summary: {counters}")
    print(f"Master file: {MASTER_EXCEL_FILE}")
    print(f"Log file:    {LOG_FILE}")


if __name__ == "__main__":
    main()
