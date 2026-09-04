"""
SCRIPT 1 of 3 — Harness extraction into the shared master Excel file.

Creates (or merges into) MASTER_EXCEL_FILE, one row per Harness service.
This script only ever writes to the columns it owns (see OWNED_COLUMNS
below) — anything Script 2 or Script 3 has already filled in for a given
service is left exactly as-is on re-run.

Credentials come from config.json in this same folder (see harness_config.py)
— fill that in once and every script in this set reads from it.

Run this first. Script 2 (GitLab extraction) and Script 3 (Harness update)
both read/write the same master Excel file afterwards.

pip install requests pyyaml openpyxl --break-system-packages
"""

import os
import time
import logging
from datetime import datetime

import yaml
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from harness_config import load_config, require_harness

# ===========================================================================
# CONFIG — credentials come from config.json (same folder as this script).
# Only things that are specific to this script stay here.
# ===========================================================================

_config = load_config()
_harness = require_harness(_config)

HARNESS_BASE_URL = _harness["base_url"]
HARNESS_API_KEY = _harness["api_key"]
HARNESS_ACCOUNT_ID = _harness["account_id"]
HARNESS_ORG_ID = _harness["org_id"]
HARNESS_PROJECT_ID = _harness["project_id"]

# Best-effort Harness NG service URL. Verify against one real service URL
# from your instance and adjust if it doesn't match — the exact routing
# has changed between Harness versions before.
HARNESS_APP_URL_TEMPLATE = (
    "https://app.harness.io/ng/account/{account}/module/cd/orgs/{org}"
    "/projects/{project}/services/{service}/summary"
)

MASTER_EXCEL_FILE = "harness_gitlab_master.xlsx"
LOG_FILE = f"script1_harness_extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

SLEEP_BETWEEN_SERVICES = 0.15
MAX_RETRIES = 3
RETRY_BACKOFF_SECONDS = 2

# ===========================================================================
# Shared column schema — every script in this 3-script set uses this exact
# list and order. Keep it identical across all three files.
# ===========================================================================

ALL_COLUMNS = [
    "Service Identifier", "Service Name", "Service Url",
    "Corresponding Gitlab component", "GitLab File Path",
    "Existing_Branch", "New_Branch",
    "Existing_ssc_appname", "New_ssc_appname",
    "Existing_ssc_appversion", "New_ssc_appversion",
    "Artifact Path",
    "Approved for Update (Y/N)", "Variables Updated",
    "Comments", "Last Run Timestamp",
]

# Columns THIS script is allowed to write. Everything else on a row is
# left untouched, even if the row already exists in the file.
OWNED_COLUMNS = [
    "Service Identifier", "Service Name", "Service Url",
    "Existing_Branch", "Existing_ssc_appname", "Existing_ssc_appversion",
    "Artifact Path",
]

# ===========================================================================
# Logging
# ===========================================================================

logger = logging.getLogger("script1")
logger.setLevel(logging.DEBUG)
fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
logger.addHandler(fh)
logger.addHandler(ch)

# ===========================================================================
# Harness API helpers
# ===========================================================================

HEADERS = {"x-api-key": HARNESS_API_KEY, "Content-Type": "application/json"}
COMMON_PARAMS = {
    "accountIdentifier": HARNESS_ACCOUNT_ID,
    "orgIdentifier": HARNESS_ORG_ID,
    "projectIdentifier": HARNESS_PROJECT_ID,
}


def request_with_retry(method: str, url: str, **kwargs) -> requests.Response:
    last_exc = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.request(method, url, timeout=30, **kwargs)
            if resp.status_code >= 500:
                raise requests.HTTPError(f"{resp.status_code} server error")
            return resp
        except requests.RequestException as exc:
            last_exc = exc
            logger.warning("Request failed (attempt %s/%s) %s %s: %s",
                            attempt, MAX_RETRIES, method, url, exc)
            time.sleep(RETRY_BACKOFF_SECONDS * attempt)
    raise last_exc


def list_all_services() -> list[dict]:
    services = []
    page, size = 0, 100
    while True:
        params = {**COMMON_PARAMS, "page": page, "size": size}
        resp = request_with_retry("GET", f"{HARNESS_BASE_URL}/ng/api/servicesV2",
                                   headers=HEADERS, params=params)
        resp.raise_for_status()
        content = resp.json()["data"]["content"]
        if not content:
            break
        for item in content:
            svc = item["service"]
            services.append({"identifier": svc["identifier"], "name": svc["name"]})
        if len(content) < size:
            break
        page += 1
    logger.info("Fetched %s services", len(services))
    return services


def get_service_yaml(service_identifier: str) -> dict:
    url = f"{HARNESS_BASE_URL}/ng/api/servicesV2/{service_identifier}"
    resp = request_with_retry("GET", url, headers=HEADERS, params=COMMON_PARAMS)
    resp.raise_for_status()
    raw_yaml = resp.json()["data"]["service"]["yaml"]
    return yaml.safe_load(raw_yaml) or {}


def get_spec(yaml_dict: dict) -> dict:
    return (yaml_dict.get("service", {}).get("serviceDefinition", {}).get("spec", {})) or {}


def extract_variables(spec: dict) -> dict:
    return {v.get("name"): v.get("value") for v in (spec.get("variables") or [])}


def extract_first_branch(spec: dict) -> str:
    """
    First manifest's branch (or commitId if it's pinned to a commit).

    Some manifests use the Harness File Store instead of Git/GitHub/
    GitLab/Bitbucket — there's no branch concept for those. Rather than
    leaving the cell blank (which reads like missing data), report the
    literal "Harness" so it's clear this manifest isn't Git-sourced at all.
    """
    manifests = spec.get("manifests", []) or []
    saw_harness_store = False
    for entry in manifests:
        m_spec = (entry.get("manifest", {}) or {}).get("spec", {}) or {}
        store = m_spec.get("store", {}) or {}
        store_spec = store.get("spec", {}) or {}
        branch = store_spec.get("branch") or store_spec.get("commitId")
        if branch:
            return branch
        if store.get("type") == "Harness":
            saw_harness_store = True
    return "Harness" if saw_harness_store else ""


ARTIFACT_PATH_FIELDS = ["artifactPath", "imagePath", "artifactDirectory", "repositoryUrl"]


def extract_first_artifact_path(spec: dict) -> str:
    artifacts = spec.get("artifacts", {}) or {}
    primary = artifacts.get("primary", {}) or {}
    sources = primary.get("sources") or ([primary] if primary.get("spec") else [])
    for src in sources:
        a_spec = src.get("spec", {}) or {}
        for field in ARTIFACT_PATH_FIELDS:
            if a_spec.get(field):
                return a_spec[field]
    return ""


# ===========================================================================
# Master Excel file — load/merge/save
# ===========================================================================


def load_or_create_master():
    if os.path.exists(MASTER_EXCEL_FILE):
        wb = load_workbook(MASTER_EXCEL_FILE)
        ws = wb.active
        header = [c.value for c in ws[1]]
        if header != ALL_COLUMNS:
            raise ValueError(
                f"{MASTER_EXCEL_FILE} exists but its header doesn't match the "
                f"expected schema. Fix the header or rename the old file before "
                f"re-running.\nExpected: {ALL_COLUMNS}\nFound:    {header}"
            )
        logger.info("Loaded existing master file with %s existing rows", ws.max_row - 1)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Master"
        ws.append(ALL_COLUMNS)
        style_header(ws)
        logger.info("Created new master file: %s", MASTER_EXCEL_FILE)
    return wb, ws


def style_header(ws):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(ALL_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(ALL_COLUMNS[col_idx - 1]) + 2, 20)
    ws.freeze_panes = "A2"


def build_row_index(ws) -> dict:
    """service_identifier -> row number, for existing rows."""
    id_col = ALL_COLUMNS.index("Service Identifier") + 1
    index = {}
    for row_idx in range(2, ws.max_row + 1):
        sid = ws.cell(row=row_idx, column=id_col).value
        if sid:
            index[sid] = row_idx
    return index


def set_owned_cell(ws, row_idx: int, column_name: str, value):
    if column_name not in OWNED_COLUMNS:
        raise ValueError(f"Script 1 tried to write a column it doesn't own: {column_name}")
    col_idx = ALL_COLUMNS.index(column_name) + 1
    ws.cell(row=row_idx, column=col_idx, value=value)


def append_comment(ws, row_idx: int, text: str):
    col_idx = ALL_COLUMNS.index("Comments") + 1
    cell = ws.cell(row=row_idx, column=col_idx)
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    new_note = f"[Script1 {stamp}] {text}"
    cell.value = f"{cell.value}\n{new_note}" if cell.value else new_note


def set_timestamp(ws, row_idx: int):
    col_idx = ALL_COLUMNS.index("Last Run Timestamp") + 1
    ws.cell(row=row_idx, column=col_idx, value=datetime.now().isoformat(timespec="seconds"))


# ===========================================================================
# Main
# ===========================================================================


def main():
    logger.info("=== Script 1 (Harness extract) started ===")
    services = list_all_services()
    wb, ws = load_or_create_master()
    row_index = build_row_index(ws)

    for i, service in enumerate(services, start=1):
        svc_id, svc_name = service["identifier"], service["name"]
        logger.info("[%s/%s] %s (%s)", i, len(services), svc_name, svc_id)

        row_idx = row_index.get(svc_id)
        if row_idx is None:
            ws.append([""] * len(ALL_COLUMNS))
            row_idx = ws.max_row
            set_owned_cell(ws, row_idx, "Service Identifier", svc_id)
            row_index[svc_id] = row_idx

        try:
            yaml_dict = get_service_yaml(svc_id)
            spec = get_spec(yaml_dict)
            variables = extract_variables(spec)
            branch = extract_first_branch(spec)
            artifact_path = extract_first_artifact_path(spec)
            service_url = HARNESS_APP_URL_TEMPLATE.format(
                account=HARNESS_ACCOUNT_ID, org=HARNESS_ORG_ID,
                project=HARNESS_PROJECT_ID, service=svc_id,
            )

            set_owned_cell(ws, row_idx, "Service Name", svc_name)
            set_owned_cell(ws, row_idx, "Service Url", service_url)
            set_owned_cell(ws, row_idx, "Existing_Branch", branch)
            set_owned_cell(ws, row_idx, "Existing_ssc_appname", variables.get("ssc_appname", ""))
            set_owned_cell(ws, row_idx, "Existing_ssc_appversion", variables.get("ssc_appversion", ""))
            set_owned_cell(ws, row_idx, "Artifact Path", artifact_path)

            if not branch:
                append_comment(ws, row_idx, "No manifest branch found")
            if not artifact_path:
                append_comment(ws, row_idx, "No artifact path found")
            if "ssc_appname" not in variables or "ssc_appversion" not in variables:
                append_comment(ws, row_idx, "One or both ssc_ variables missing on Harness")

        except Exception as exc:  # noqa: BLE001 — keep going across 330 services
            logger.error("Failed on service '%s': %s", svc_name, exc)
            set_owned_cell(ws, row_idx, "Service Name", svc_name)
            append_comment(ws, row_idx, f"ERROR during extraction: {exc}")

        set_timestamp(ws, row_idx)
        time.sleep(SLEEP_BETWEEN_SERVICES)

    wb.save(MASTER_EXCEL_FILE)
    logger.info("=== Script 1 complete: %s services processed, saved to %s ===",
                len(services), MASTER_EXCEL_FILE)
    print(f"\nDone. {len(services)} services processed.")
    print(f"Master file: {MASTER_EXCEL_FILE}")
    print(f"Log file:    {LOG_FILE}")


if __name__ == "__main__":
    main()
